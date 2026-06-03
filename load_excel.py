#!/usr/bin/env python3
"""
Carga histórico de partes diarios desde Excel a Supabase via SQL batches.
Genera SQL que se puede ejecutar via MCP execute_sql.
"""
import openpyxl
import json
import re
from datetime import date, datetime

EXCEL_PATH = "/Users/natalia/Documents/Yura maquina/dashboard-planta-yura/02. Reporte Despacho Dia 2025v1 (004).xlsx"

SKIP_SHEETS = {'Plan ventas 2026', 'DESPACHO POR MATERIAL'}

# tipo_cemento from DB (id -> nombre)
TIPO_CEMENTO = {
    1: "CEM. FRONTERA GU X 42.5 KG",
    2: "CEMENTO FRONTERA IP x 42.5 KG",
    3: "CEMENTO HS+R X 42.5 KG RUMI",
    4: "CEMENTO PORTLAND TIPO I X 42.5 KG YURA",
    5: "CEMENTO YURA ANTISALITRE (MS) X 42.5 KG",
    6: "CEMENTO YURA MAX (HS) X 42.5KG",
    7: "CEMENTO YURA PRO (HE) 42.5KG",
    8: "CEMENTO YURA PRO EXPORT BRASIL x 42.5 kg",
    9: "CEMENTO BLANCO X 25 KG YURA NACIONAL",
    10: "CEMENTO YURA PRO (HE) BBx1.5TM INC.ENV",
    11: "CEM ALT REST IN HE BB1.5TM YURA ENV CHIL",
    12: "CEM. PORT. PUZ. IP BBx1.5TM INC.ENV YURA",
    13: "CEM.YURA ANTISALITRE MS BBx1.5TM INC.ENV",
    14: "CEMENTO YURA MAX BBx1.5TM INC.ENV",
    15: "CEM. PORT. I BBx1.5TM INC.ENV YURA",
    16: "CEMENTO YURA PRO (HE) GRANEL EN BOMBONA",
    17: "CEMENTO YURA IP A GRANEL EN BOMBONA",
    18: "CEMENTO YURA MAX GRANEL BOMBONA",
}

# Normalize: uppercase, collapse spaces
def norm(s):
    if not s:
        return ""
    return re.sub(r'\s+', ' ', str(s).strip().upper())

# Build reverse map: normalized name -> id
NOMBRE_TO_ID = {}
for tid, nombre in TIPO_CEMENTO.items():
    NOMBRE_TO_ID[norm(nombre)] = tid

# Additional aliases for common variants in Excel
ALIASES = {
    "CEMENTO YURA MAX  (HS) X 42.5KG": 6,
    "CEMENTO YURA MAX (HS) X 42.5KG": 6,
    "CEMENTO YURA MAX  BBx1.5TM INC.ENV": 14,
    "CEM. PORT. PUZ. IP BBx1.5TM INC.ENV YURA": 12,
    "CEM.ANTISALITRE HS BBx1.5TM INC.ENV YURA": 13,
    "CEM.YURA ANTISALITRE MS BBx1.5TM INC.ENV": 13,
    "CEMENTO ANTISALITRE TIPO HS X 25KG YURA": None,  # not in DB
    "CEM.PORTL.IP(GRADO CORRIENTE)X25KG CHILE": None,  # not in DB
    "CEMENTO BLANCO BIGBAG X 1.5TM EXP.BOLIVIA": None,
    "FILLER INDUSTRIAL BB1.5TM": None,
    "CEM YURA ESPECIAL (HS) X 42.5 KG": None,
    "CEMENTO PORT. PUZ. IP X 25 KG YURA": None,
    "CEMENTO PARA ACABADOS TIPO MH  25KG YURA": None,
    "CEMENTO PARA ACABADOS TIPO MH 25KG YURA": None,
    "CEMENTO BLANCO BIGBAG X 1.5TM EXP.BOLIVIA": None,
    "CEM YURA ESPECIAL  (HS) X 42.5 KG": None,
    "CEM.ANTISALITRE HS BBX1.5TM INC.ENV YURA": 13,
    "CEM. FRONTERA GU X 42.5KG": 1,
}
for k, v in ALIASES.items():
    if v is not None:
        NOMBRE_TO_ID[norm(k)] = v

def lookup_tipo(name):
    n = norm(name)
    if n in NOMBRE_TO_ID:
        return NOMBRE_TO_ID[n]
    # Try partial matching for common patterns
    return None

def to_num(v, default=0):
    if v is None:
        return default
    try:
        f = float(v)
        return f if f == f else default  # nan check
    except (TypeError, ValueError):
        return default

def to_int(v, default=0):
    return int(to_num(v, default))

def esc(s):
    """Escape single quotes for SQL."""
    if s is None:
        return ''
    return str(s).replace("'", "''")

MONTH_NAMES = {
    'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4, 'MAYO': 5, 'JUNIO': 6,
    'JULIO': 7, 'AGOSTO': 8, 'SETIEMBRE': 9, 'SEPTIEMBRE': 9, 'OCTUBRE': 10,
    'NOVIEMBRE': 11, 'DICIEMBRE': 12
}

def parse_sheet_date(sheet_name, rows):
    """Try to find date from cell, else derive from sheet name."""
    r0 = rows[0] if rows else ()
    name = sheet_name.strip()

    # For numeric-only sheet names (day only)
    m_num = re.match(r'^(\d{1,2})$', name)
    if m_num:
        day_from_name = int(m_num.group(1))
        # Try to get date from cell - validate day matches sheet name
        for col_idx in [1, 7]:
            if col_idx < len(r0) and isinstance(r0[col_idx], datetime):
                cell_date = r0[col_idx].date()
                if cell_date.day == day_from_name and cell_date.year in (2025, 2026):
                    return cell_date
        # Cell date invalid - derive from "ME XXX" in row0
        for cell in r0:
            if isinstance(cell, str):
                cu = cell.strip().upper()
                if cu.startswith('ME '):
                    mes_str = cu[3:].strip()
                    if mes_str in MONTH_NAMES:
                        mo = MONTH_NAMES[mes_str]
                        year = 2025
                        try:
                            return date(year, mo, day_from_name)
                        except ValueError:
                            pass
        return None

    # For dd.mm.yy format
    m = re.match(r'^(\d{1,2})\.(\d{2})\.(\d{2})$', name)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = 2000 + y
        # Prefer cell if it matches
        for col_idx in [1, 7]:
            if col_idx < len(r0) and isinstance(r0[col_idx], datetime):
                cell_date = r0[col_idx].date()
                if cell_date.day == d and cell_date.month == mo and cell_date.year == year:
                    return cell_date
        try:
            return date(year, mo, d)
        except ValueError:
            return None

    # For dd.mm format (year = 2025)
    m = re.match(r'^(\d{1,2})\.(\d{2})$', name)
    if m:
        d, mo = int(m.group(1)), int(m.group(2))
        # Prefer cell if it matches
        for col_idx in [1, 7]:
            if col_idx < len(r0) and isinstance(r0[col_idx], datetime):
                cell_date = r0[col_idx].date()
                if cell_date.day == d and cell_date.month == mo and cell_date.year in (2025, 2026):
                    return cell_date
        year = 2025
        try:
            return date(year, mo, d)
        except ValueError:
            return None

    return None

def find_row_with(rows, keyword, col=1):
    """Find row index where column col contains keyword."""
    kw = norm(keyword)
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell is not None and norm(str(cell)) == kw:
                return i, j
    return None, None

def find_row_containing(rows, keyword):
    """Find row index where any cell contains keyword."""
    kw = norm(keyword)
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if cell is not None and kw in norm(str(cell)):
                return i, j
    return None, None

def parse_sheet(sheet_name, ws):
    rows = list(ws.iter_rows(values_only=True))

    result = {
        'fecha': None,
        'veh_llamado': 0, 'veh_proceso': 0, 'veh_playa': 0,
        'acumulado_ajuste': 0, 'comentario': '',
        'despacho': [],  # list of (tipo_id, bolsas, tm)
        'maquinas': [],  # list of (maquina_id, horas, ratio_ecs, operativos, comentario)
        'temporales': [],  # list of (temporal_id, inventario)
        'compuertas': [],  # list of (numero, horas, comentario)
        'unmapped_tipos': [],
    }

    result['fecha'] = parse_sheet_date(sheet_name, rows)
    if result['fecha'] is None:
        return None, f"No se pudo determinar fecha"

    # Find DESPACHO DEL DÍA POR TIPO DE CEMENTO
    desp_row, desp_col = find_row_containing(rows, 'DESPACHO DEL DÍA POR TIPO DE CEMENTO')
    if desp_row is None:
        desp_row, desp_col = find_row_containing(rows, 'DESPACHO DEL DIA POR TIPO DE CEMENTO')

    end_desp_row = None
    if desp_row is not None:
        # Find column offsets: nombre is at desp_col (usually 1), bolsas col+1, tm col+2
        name_col = desp_col if desp_col is not None else 1
        bolsas_col = name_col + 1
        tm_col = name_col + 2

        for i in range(desp_row + 1, min(desp_row + 50, len(rows))):
            row = rows[i]
            if not row or len(row) <= name_col:
                continue
            cell_name = row[name_col] if name_col < len(row) else None
            if cell_name is None:
                continue
            cell_name_str = norm(str(cell_name))
            if 'DESPACHO T.M. DEL DIA' in cell_name_str or 'DESPACHO T.M. DEL DÍA' in cell_name_str:
                end_desp_row = i
                # Get acumulado from tm_col
                if tm_col < len(row) and row[tm_col] is not None:
                    result['acumulado_ajuste'] = to_num(row[tm_col])
                break
            if isinstance(cell_name, str) and cell_name.strip():
                tipo_id = lookup_tipo(cell_name)
                bolsas = to_num(row[bolsas_col] if bolsas_col < len(row) else None)
                tm = to_num(row[tm_col] if tm_col < len(row) else None)
                if tipo_id is not None:
                    if bolsas != 0 or tm != 0:
                        result['despacho'].append((tipo_id, bolsas, tm))
                else:
                    n = norm(cell_name)
                    if n and n not in ('ACUMULADO MES', 'DESPACHO T.M. DEL DIA'):
                        result['unmapped_tipos'].append(cell_name.strip())

    # Find TEMPORALES
    temp_row, _ = find_row_containing(rows, 'TEMPORALES')
    if temp_row is not None:
        temp_map = {'TEMPORAL 1': 1, 'TEMPORAL 2': 2, 'MISTI': 3}
        for i in range(temp_row + 1, min(temp_row + 10, len(rows))):
            row = rows[i]
            if not row:
                continue
            for col in range(len(row)):
                cell = row[col]
                if cell is not None and norm(str(cell)) in [norm(k) for k in temp_map.keys()]:
                    label = norm(str(cell))
                    for k, tid in temp_map.items():
                        if norm(k) == label:
                            inv_col = col + 1
                            inv = to_num(row[inv_col] if inv_col < len(row) else None)
                            result['temporales'].append((tid, inv))
                            break

    # Find CARROS/PLAYA - look for row with LLAMADO, PROCESO DE C, TOTAL
    for i, row in enumerate(rows):
        row_text = [norm(str(c)) if c is not None else '' for c in row]
        if 'LLAMADO' in row_text and 'PROCESO DE C' in row_text:
            llamado_col = row_text.index('LLAMADO')
            proceso_col = row_text.index('PROCESO DE C')
            # Next row has numbers
            if i + 1 < len(rows):
                next_row = rows[i + 1]
                # playa is col before llamado
                playa_col = llamado_col - 1
                result['veh_playa'] = to_int(next_row[playa_col] if playa_col >= 0 and playa_col < len(next_row) else None)
                result['veh_llamado'] = to_int(next_row[llamado_col] if llamado_col < len(next_row) else None)
                result['veh_proceso'] = to_int(next_row[proceso_col] if proceso_col < len(next_row) else None)
            break

    # Find maquinas: look for H1,H2... rows
    # They appear after a row with HORAS MAQUINA or just find rows starting with H1..H5,V
    maquina_ids = {'H1', 'H2', 'H3', 'H4', 'H5', 'V'}
    horas_col = 1  # default
    ratio_col = 2
    operativos_col = 5  # varies

    # Try to find header row with HORAS MAQUINA
    hm_row, hm_col = find_row_containing(rows, 'HORAS MAQUINA')
    if hm_row is not None:
        header = rows[hm_row]
        # Find column positions
        for j, cell in enumerate(header):
            cn = norm(str(cell)) if cell else ''
            if 'HORAS MAQUINA' in cn:
                horas_col = j
            elif 'RATIO ECS' in cn:
                ratio_col = j
            elif 'OPERATIVOS' in cn:
                operativos_col = j
        # Next rows are machines
        for i in range(hm_row + 1, min(hm_row + 10, len(rows))):
            row = rows[i]
            if not row:
                continue
            mid = str(row[0]).strip() if row[0] is not None else ''
            if mid in maquina_ids:
                horas = to_num(row[horas_col] if horas_col < len(row) else None)
                ratio = to_int(row[ratio_col] if ratio_col < len(row) else None)
                ops = to_int(row[operativos_col] if operativos_col < len(row) else None)
                comentario = ''
                # check if horas is a string (comment)
                raw_horas = row[horas_col] if horas_col < len(row) else None
                if isinstance(raw_horas, str):
                    comentario = raw_horas.strip()
                    horas = 0
                result['maquinas'].append((mid, horas, ratio, ops, comentario))
    else:
        # Fallback: find rows where col0 is H1..H5,V
        for i, row in enumerate(rows):
            if not row:
                continue
            mid = str(row[0]).strip() if row[0] is not None else ''
            if mid in maquina_ids:
                raw_horas = row[1] if len(row) > 1 else None
                comentario = ''
                if isinstance(raw_horas, str):
                    comentario = raw_horas.strip()
                    horas = 0
                else:
                    horas = to_num(raw_horas)
                ratio = to_int(row[2] if len(row) > 2 else None)
                # Find operativos - try col 5 and 6
                ops = 0
                for op_col in [5, 6]:
                    if op_col < len(row) and isinstance(row[op_col], (int, float)):
                        ops = to_int(row[op_col])
                        break
                result['maquinas'].append((mid, horas, ratio, ops, comentario))

    # Find compuertas - under SILO 8 / COMPUERTA / HR
    silo_row, _ = find_row_containing(rows, 'SILO 8')
    if silo_row is not None:
        # Find column with COMPUERTA label nearby
        comp_col = 2  # default
        hr_col = 3
        # Look at row silo_row and nearby for COMPUERTA header
        for check_row in range(max(0, silo_row - 2), min(silo_row + 3, len(rows))):
            row = rows[check_row]
            for j, cell in enumerate(row):
                if cell is not None and 'COMPUERTA' in norm(str(cell)):
                    comp_col = j
                elif cell is not None and norm(str(cell)) == 'HR':
                    hr_col = j

        # Parse compuerta rows starting from silo_row
        for i in range(silo_row, min(silo_row + 15, len(rows))):
            row = rows[i]
            if not row or comp_col >= len(row):
                continue
            num_cell = row[comp_col]
            hr_cell = row[hr_col] if hr_col < len(row) else None
            # num must be integer 1-8
            if isinstance(num_cell, (int, float)) and 1 <= num_cell <= 8:
                numero = int(num_cell)
                horas_c = to_num(hr_cell)
                # comentario could be string in hr_col
                comentario_c = ''
                if isinstance(hr_cell, str):
                    comentario_c = hr_cell.strip()
                    horas_c = 0
                result['compuertas'].append((numero, horas_c, comentario_c))

    return result, None


def generate_sql_for_sheet(data):
    """Generate SQL statements for one sheet's data."""
    sqls = []
    fecha = data['fecha']
    fecha_str = fecha.isoformat()

    # Upsert parte_diario
    sqls.append(f"""
INSERT INTO parte_diario (fecha, veh_llamado, veh_proceso, veh_playa, acumulado_ajuste, comentario)
VALUES ('{fecha_str}', {data['veh_llamado']}, {data['veh_proceso']}, {data['veh_playa']}, {data['acumulado_ajuste']}, '{esc(data['comentario'])}')
ON CONFLICT (fecha) DO UPDATE SET
  veh_llamado = EXCLUDED.veh_llamado,
  veh_proceso = EXCLUDED.veh_proceso,
  veh_playa = EXCLUDED.veh_playa,
  acumulado_ajuste = EXCLUDED.acumulado_ajuste,
  comentario = EXCLUDED.comentario
RETURNING id;
""".strip())

    return sqls


def main():
    print("Cargando Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheets = wb.sheetnames

    results = []
    skipped = []
    unmapped_global = set()
    errors = []

    for sheet_name in sheets:
        # Skip aux sheets
        if sheet_name in SKIP_SHEETS:
            skipped.append(f"{sheet_name} (auxiliar)")
            continue
        if re.match(r'^Hoja\d+$', sheet_name, re.IGNORECASE):
            skipped.append(f"{sheet_name} (HojaN)")
            continue

        ws = wb[sheet_name]
        try:
            data, err = parse_sheet(sheet_name, ws)
        except Exception as e:
            errors.append(f"{sheet_name}: {e}")
            skipped.append(f"{sheet_name} (error: {e})")
            continue

        if data is None:
            skipped.append(f"{sheet_name} ({err})")
            continue

        for u in data['unmapped_tipos']:
            unmapped_global.add(u)

        results.append((sheet_name, data))

    print(f"\nHojas parseadas: {len(results)}")
    print(f"Hojas saltadas: {len(skipped)}")
    print(f"Errores: {len(errors)}")

    # Output structured data as JSON for the loader
    output = []
    for sheet_name, data in results:
        output.append({
            'sheet': sheet_name,
            'fecha': data['fecha'].isoformat(),
            'veh_llamado': data['veh_llamado'],
            'veh_proceso': data['veh_proceso'],
            'veh_playa': data['veh_playa'],
            'acumulado_ajuste': data['acumulado_ajuste'],
            'comentario': data['comentario'],
            'despacho': [{'tipo_id': t, 'bolsas': b, 'tm': tm} for t, b, tm in data['despacho']],
            'maquinas': [{'maquina_id': m, 'horas': h, 'ratio_ecs': r, 'operativos': o, 'comentario': c} for m, h, r, o, c in data['maquinas']],
            'temporales': [{'temporal_id': tid, 'inventario': inv} for tid, inv in data['temporales']],
            'compuertas': [{'numero': n, 'horas': h, 'comentario': c} for n, h, c in data['compuertas']],
        })

    with open('/tmp/parsed_partes.json', 'w') as f:
        json.dump(output, f, indent=2)

    print(f"\nDatos guardados en /tmp/parsed_partes.json")
    print(f"\nHojas saltadas:")
    for s in skipped:
        print(f"  - {s}")

    print(f"\nTipos de cemento no mapeados ({len(unmapped_global)}):")
    for u in sorted(unmapped_global):
        print(f"  - {u}")

    return output, skipped, unmapped_global

if __name__ == '__main__':
    main()
